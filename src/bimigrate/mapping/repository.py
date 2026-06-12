"""Feature Mapping Repository: SQLite store + Excel/JSON export.

Every conversion decision in the pipeline cites a FeatureMapping record
by feature_name (traceability requirement, Section 5).
"""

from __future__ import annotations

import json
import sqlite3
from pathlib import Path

from bimigrate.models.core import FeatureMapping, tier_for_confidence

_SCHEMA = """
CREATE TABLE IF NOT EXISTS feature_mappings (
    feature_name TEXT NOT NULL,
    source_tool TEXT NOT NULL,
    source_layer TEXT NOT NULL,
    powerbi_equivalent TEXT,
    migration_complexity TEXT NOT NULL,
    automation_possibility TEXT NOT NULL,
    fallback_strategy TEXT NOT NULL,
    confidence_score REAL NOT NULL,
    notes TEXT,
    example_source TEXT,
    example_target TEXT,
    PRIMARY KEY (source_tool, feature_name)
);
"""


class MappingRepository:
    def __init__(self, db_path: Path | str = ":memory:"):
        self.db_path = str(db_path)
        self._conn = sqlite3.connect(self.db_path)
        self._conn.row_factory = sqlite3.Row
        self._conn.executescript(_SCHEMA)

    def close(self) -> None:
        self._conn.close()

    def init_from_seeds(self) -> int:
        from bimigrate.mapping.seed_data import ALL_SEED_MAPPINGS

        count = 0
        for tool, rows in ALL_SEED_MAPPINGS.items():
            for name, layer, equiv, complexity, automation, fallback, conf, notes, ex_src, ex_tgt in rows:
                mapping = FeatureMapping(
                    feature_name=name,
                    source_tool=tool,
                    source_layer=layer,
                    powerbi_equivalent=equiv,
                    migration_complexity=complexity,
                    automation_possibility=automation,
                    fallback_strategy=fallback or "",
                    confidence_score=conf,
                    notes=notes or "",
                    example_source=ex_src,
                    example_target=ex_tgt,
                )
                self.upsert(mapping)
                count += 1
        return count

    def upsert(self, m: FeatureMapping) -> None:
        self._conn.execute(
            "INSERT OR REPLACE INTO feature_mappings VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (
                m.feature_name,
                m.source_tool,
                m.source_layer,
                m.powerbi_equivalent,
                m.migration_complexity,
                m.automation_possibility,
                m.fallback_strategy,
                m.confidence_score,
                m.notes,
                m.example_source,
                m.example_target,
            ),
        )
        self._conn.commit()

    def get(self, source_tool: str, feature_name: str) -> FeatureMapping | None:
        row = self._conn.execute(
            "SELECT * FROM feature_mappings WHERE source_tool=? AND feature_name=?",
            (source_tool, feature_name),
        ).fetchone()
        return self._row_to_model(row) if row else None

    def for_tool(self, source_tool: str) -> list[FeatureMapping]:
        rows = self._conn.execute(
            "SELECT * FROM feature_mappings WHERE source_tool=? ORDER BY source_layer, feature_name",
            (source_tool,),
        ).fetchall()
        return [self._row_to_model(r) for r in rows]

    def all(self) -> list[FeatureMapping]:
        rows = self._conn.execute(
            "SELECT * FROM feature_mappings ORDER BY source_tool, source_layer, feature_name"
        ).fetchall()
        return [self._row_to_model(r) for r in rows]

    def unsupported(self, source_tool: str | None = None) -> list[FeatureMapping]:
        """Features with no native Power BI equivalent — these drive the
        'unsupported features + recommendation' section of reports."""
        return [
            m
            for m in (self.for_tool(source_tool) if source_tool else self.all())
            if m.powerbi_equivalent is None or m.automation_possibility == "manual"
        ]

    @staticmethod
    def _row_to_model(row: sqlite3.Row) -> FeatureMapping:
        return FeatureMapping(
            feature_name=row["feature_name"],
            source_tool=row["source_tool"],
            source_layer=row["source_layer"],
            powerbi_equivalent=row["powerbi_equivalent"],
            migration_complexity=row["migration_complexity"],
            automation_possibility=row["automation_possibility"],
            fallback_strategy=row["fallback_strategy"],
            confidence_score=row["confidence_score"],
            notes=row["notes"] or "",
            example_source=row["example_source"],
            example_target=row["example_target"],
        )

    # -- exports ---------------------------------------------------------

    def export_json(self, path: Path) -> Path:
        path.parent.mkdir(parents=True, exist_ok=True)
        rows = [m.model_dump() | {"tier": tier_for_confidence(m.confidence_score).value} for m in self.all()]
        path.write_text(json.dumps(rows, indent=2), encoding="utf-8")
        return path

    def export_excel(self, path: Path) -> Path:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)
        tier_fill = {
            "AUTO": PatternFill("solid", fgColor="C6EFCE"),
            "ASSISTED": PatternFill("solid", fgColor="FFEB9C"),
            "MANUAL": PatternFill("solid", fgColor="FFC7CE"),
        }
        headers = [
            "Feature",
            "Layer",
            "Power BI equivalent",
            "Complexity",
            "Automation",
            "Tier",
            "Confidence",
            "Fallback strategy",
            "Notes",
        ]
        for tool in ("tableau", "spotfire", "qlikview", "qliksense"):
            mappings = self.for_tool(tool)
            if not mappings:
                continue
            ws = wb.create_sheet(tool.capitalize())
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for m in mappings:
                tier = tier_for_confidence(m.confidence_score).value
                ws.append(
                    [
                        m.feature_name,
                        m.source_layer,
                        m.powerbi_equivalent or "(no native equivalent)",
                        m.migration_complexity,
                        m.automation_possibility,
                        tier,
                        m.confidence_score,
                        m.fallback_strategy,
                        m.notes,
                    ]
                )
                ws.cell(row=ws.max_row, column=6).fill = tier_fill[tier]
            for col, width in zip("ABCDEFGHI", (38, 14, 45, 14, 12, 10, 11, 60, 50)):
                ws.column_dimensions[col].width = width
        wb.save(path)
        return path
