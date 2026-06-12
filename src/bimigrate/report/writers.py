"""Inventory / assessment report writers: Excel, HTML, CSV, JSON."""

from __future__ import annotations

import csv
import html
import json
from pathlib import Path

from bimigrate.engine.complexity import effort_hours
from bimigrate.engine.dedup import DuplicateCluster, dedup_savings_estimate
from bimigrate.models.inventory import ArtifactInventory


def write_reports(
    inventories: list[ArtifactInventory],
    clusters: list[DuplicateCluster],
    out_dir: Path,
    errors: list[tuple[str, str]] | None = None,
) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    written = [
        _write_json(inventories, clusters, out_dir),
        _write_csv(inventories, out_dir),
        _write_excel(inventories, clusters, out_dir, errors or []),
        _write_html(inventories, clusters, out_dir, errors or []),
    ]
    return written


def _rows(inventories: list[ArtifactInventory]) -> list[dict]:
    return [
        {
            "artifact": inv.artifact_name,
            "path": inv.artifact_path,
            "source_tool": inv.source_tool.value,
            "complexity_score": inv.complexity_score,
            "complexity_band": inv.complexity_band.value,
            "estimated_effort_hours": effort_hours(inv),
            "worksheets": len(inv.worksheets),
            "visuals": inv.visual_count,
            "dashboards": len(inv.dashboards),
            "calculations": len(inv.calculations),
            "connections": len(inv.connections),
            "scripts": len(inv.scripts),
            "security_rules": len(inv.security),
            "parse_issues": len(inv.issues),
            "fingerprint": inv.fingerprint,
        }
        for inv in sorted(inventories, key=lambda i: -i.complexity_score)
    ]


def _write_json(inventories, clusters, out_dir: Path) -> Path:
    path = out_dir / "inventory.json"
    payload = {
        "summary": _summary(inventories, clusters),
        "artifacts": [inv.model_dump(mode="json") for inv in inventories],
        "duplicate_clusters": [
            {
                "representative": c.representative,
                "members": c.members,
                "retire_candidates": c.retire_candidates,
            }
            for c in clusters
        ],
    }
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    return path


def _write_csv(inventories, out_dir: Path) -> Path:
    path = out_dir / "inventory.csv"
    rows = _rows(inventories)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0]) if rows else ["artifact"])
        writer.writeheader()
        writer.writerows(rows)
    return path


def _summary(inventories, clusters) -> dict:
    by_tool: dict[str, int] = {}
    by_band: dict[str, int] = {}
    for inv in inventories:
        by_tool[inv.source_tool.value] = by_tool.get(inv.source_tool.value, 0) + 1
        by_band[inv.complexity_band.value] = by_band.get(inv.complexity_band.value, 0) + 1
    return {
        "artifact_count": len(inventories),
        "by_source_tool": by_tool,
        "by_complexity_band": by_band,
        "total_visuals": sum(inv.visual_count for inv in inventories),
        "total_calculations": sum(len(inv.calculations) for inv in inventories),
        "duplicate_clusters": len(clusters),
        "dedup_savings": dedup_savings_estimate(clusters, inventories),
    }


def _write_excel(inventories, clusters, out_dir: Path, errors) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    path = out_dir / "inventory.xlsx"
    wb = Workbook()
    band_fill = {
        "simple": PatternFill("solid", fgColor="C6EFCE"),
        "medium": PatternFill("solid", fgColor="FFEB9C"),
        "complex": PatternFill("solid", fgColor="FFC7CE"),
    }

    ws = wb.active
    ws.title = "Summary"
    summary = _summary(inventories, clusters)
    ws.append(["Metric", "Value"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    flat = {
        "Artifacts": summary["artifact_count"],
        **{f"  {k}": v for k, v in summary["by_source_tool"].items()},
        "Visuals": summary["total_visuals"],
        "Calculations": summary["total_calculations"],
        "Near-duplicate clusters": summary["duplicate_clusters"],
        "Total effort (h)": summary["dedup_savings"]["total_effort_hours"],
        "Dedup savings (h)": summary["dedup_savings"]["saved_effort_hours"],
        "Dedup savings (%)": summary["dedup_savings"]["savings_pct"],
        "Parse errors": len(errors),
    }
    for key, value in flat.items():
        ws.append([key, value])
    ws.column_dimensions["A"].width = 32

    ws = wb.create_sheet("Artifacts")
    rows = _rows(inventories)
    if rows:
        ws.append([h.replace("_", " ").title() for h in rows[0]])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append(list(row.values()))
            ws.cell(row=ws.max_row, column=5).fill = band_fill[row["complexity_band"]]
        for col, width in zip("ABCDEFGHIJKLMNO", (28, 45, 11, 12, 12, 12, 10, 8, 10, 12, 11, 8, 12, 11, 18)):
            ws.column_dimensions[col].width = width

    ws = wb.create_sheet("Feature Flags")
    ws.append(["Artifact", "Feature", "Count"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for inv in inventories:
        for flag, count in sorted(inv.feature_flags.items()):
            if count:
                ws.append([inv.artifact_name, flag, count])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28

    ws = wb.create_sheet("Duplicates")
    ws.append(["Cluster", "Representative (migrate)", "Member (retire candidate)"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, cluster in enumerate(clusters, 1):
        for member in cluster.retire_candidates:
            ws.append([f"cluster-{i}", cluster.representative, member])
    ws.column_dimensions["B"].width = ws.column_dimensions["C"].width = 60

    if errors:
        ws = wb.create_sheet("Errors")
        ws.append(["Path", "Error"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for err_path, err in errors:
            ws.append([err_path, err])
        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 80

    wb.save(path)
    return path


def _write_html(inventories, clusters, out_dir: Path, errors) -> Path:
    path = out_dir / "inventory.html"
    summary = _summary(inventories, clusters)
    savings = summary["dedup_savings"]
    rows = _rows(inventories)
    band_color = {"simple": "#1a7f37", "medium": "#9a6700", "complex": "#cf222e"}

    table_rows = "\n".join(
        "<tr><td>{a}</td><td>{t}</td><td style='color:{c};font-weight:600'>{b}</td>"
        "<td>{s:.1f}</td><td>{e}</td><td>{v}</td><td>{calc}</td></tr>".format(
            a=html.escape(row["artifact"]),
            t=row["source_tool"],
            c=band_color[row["complexity_band"]],
            b=row["complexity_band"],
            s=row["complexity_score"],
            e=row["estimated_effort_hours"],
            v=row["visuals"],
            calc=row["calculations"],
        )
        for row in rows
    )
    doc = f"""<!doctype html><html><head><meta charset="utf-8">
<title>BI Migration Discovery Report</title>
<style>
body{{font-family:'Segoe UI',sans-serif;margin:2rem;color:#1f2328}}
.cards{{display:flex;gap:1rem;flex-wrap:wrap;margin-bottom:2rem}}
.card{{border:1px solid #d1d9e0;border-radius:8px;padding:1rem 1.5rem;min-width:140px}}
.card b{{display:block;font-size:1.6rem}}
table{{border-collapse:collapse;width:100%}}
th,td{{border-bottom:1px solid #d1d9e0;padding:.4rem .6rem;text-align:left;font-size:.9rem}}
th{{background:#f6f8fa}}
</style></head><body>
<h1>BI Migration Discovery Report</h1>
<div class="cards">
<div class="card"><b>{summary["artifact_count"]}</b>artifacts</div>
<div class="card"><b>{summary["total_visuals"]}</b>visuals</div>
<div class="card"><b>{summary["total_calculations"]}</b>calculations</div>
<div class="card"><b>{summary["duplicate_clusters"]}</b>duplicate clusters</div>
<div class="card"><b>{savings["savings_pct"]}%</b>dedup savings
({savings["saved_effort_hours"]}h of {savings["total_effort_hours"]}h)</div>
<div class="card"><b>{len(errors)}</b>parse errors</div>
</div>
<h2>Artifacts by complexity</h2>
<table><tr><th>Artifact</th><th>Tool</th><th>Band</th><th>Score</th>
<th>Effort (h)</th><th>Visuals</th><th>Calcs</th></tr>
{table_rows}
</table>
<p style="margin-top:2rem;color:#59636e">Effort estimates are assessment-mode inputs,
not commitments. Accuracy is tracked per tier and complexity band in validation
reports — never as a blended number.</p>
</body></html>"""
    path.write_text(doc, encoding="utf-8")
    return path
